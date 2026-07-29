import json
import math
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import linprog
from scipy.stats import linregress, spearmanr

ROOT = Path(__file__).resolve().parents[1]
YEARS = [1990, 2000, 2010, 2020]


def load_data():
    wb = load_workbook(ROOT / "rowdata" / "YRD_city_inputs_1990_2020.xlsx", data_only=True, read_only=True)
    ws = wb["Model Inputs"]
    headers = {cell.value: cell.column for cell in ws[1]}
    required = [
        "Province-level unit",
        "City",
        "Abbreviation",
        "Year",
        "Built-up land area (km²)",
        "Permanent resident population (10,000 persons)",
        "Real GDP (CNY 100 million, constant 1978 prices)",
        "CO₂ emissions (10,000 tonnes)",
    ]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Missing input columns: {missing}")
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        city = values[headers["City"] - 1]
        if not city:
            continue
        rows.append({
            "province": values[headers["Province-level unit"] - 1],
            "city": city,
            "short": values[headers["Abbreviation"] - 1],
            "year": int(values[headers["Year"] - 1]),
            "land": float(values[headers["Built-up land area (km²)"] - 1]),
            "population": float(values[headers["Permanent resident population (10,000 persons)"] - 1]),
            "real_gdp": float(values[headers["Real GDP (CNY 100 million, constant 1978 prices)"] - 1]),
            "co2": float(values[headers["CO₂ emissions (10,000 tonnes)"] - 1]),
        })
    assert len(rows) == 164, len(rows)
    assert sorted({row["year"] for row in rows}) == YEARS
    assert len({row["city"] for row in rows}) == 41
    return rows


def sbm_undesirable(x0, yg0, yb0, X, YG, YB, vrs=False):
    # Tone's non-oriented SBM with undesirable output, transformed via Charnes-Cooper.
    n, m = X.shape
    s1, s2 = YG.shape[1], YB.shape[1]
    # z = [Lambda(n), Sminus(m), Sgood(s1), Sbad(s2), t]
    nvar = n + m + s1 + s2 + 1
    i_lam = slice(0, n)
    i_sm = slice(n, n + m)
    i_sg = slice(n + m, n + m + s1)
    i_sb = slice(n + m + s1, n + m + s1 + s2)
    it = nvar - 1
    c = np.zeros(nvar)
    c[i_sm] = -1.0 / (m * x0)
    c[it] = 1.0

    Aeq, beq = [], []
    # X Lambda + S- - x0*t = 0
    for k in range(m):
        row = np.zeros(nvar)
        row[i_lam] = X[:, k]
        row[n + k] = 1.0
        row[it] = -x0[k]
        Aeq.append(row); beq.append(0.0)
    # Yg Lambda - Sg - yg0*t = 0
    for k in range(s1):
        row = np.zeros(nvar)
        row[i_lam] = YG[:, k]
        row[n + m + k] = -1.0
        row[it] = -yg0[k]
        Aeq.append(row); beq.append(0.0)
    # Yb Lambda + Sb - yb0*t = 0
    for k in range(s2):
        row = np.zeros(nvar)
        row[i_lam] = YB[:, k]
        row[n + m + s1 + k] = 1.0
        row[it] = -yb0[k]
        Aeq.append(row); beq.append(0.0)
    # t + average normalized output slacks = 1
    row = np.zeros(nvar)
    row[i_sg] = 1.0 / ((s1 + s2) * yg0)
    row[i_sb] = 1.0 / ((s1 + s2) * yb0)
    row[it] = 1.0
    Aeq.append(row); beq.append(1.0)
    if vrs:
        row = np.zeros(nvar)
        row[i_lam] = 1.0
        row[it] = -1.0
        Aeq.append(row); beq.append(0.0)
    res = linprog(c, A_eq=np.array(Aeq), b_eq=np.array(beq), bounds=[(0, None)] * nvar, method="highs")
    if not res.success:
        raise RuntimeError(res.message)
    return float(res.fun)


def directional_distance(x0, yg0, yb0, X, YG, YB, vrs=False):
    """Output-oriented directional distance with direction (yg0, -yb0)."""
    n = X.shape[0]
    # z = [lambda_1...lambda_n, beta], maximize beta -> minimize -beta
    c = np.zeros(n + 1)
    c[-1] = -1.0
    Aub, bub = [], []
    # X lambda <= x0
    for k in range(X.shape[1]):
        row = np.zeros(n + 1); row[:n] = X[:, k]
        Aub.append(row); bub.append(x0[k])
    # -Yg lambda + beta*yg0 <= -yg0
    for k in range(YG.shape[1]):
        row = np.zeros(n + 1); row[:n] = -YG[:, k]; row[-1] = yg0[k]
        Aub.append(row); bub.append(-yg0[k])
    # Yb lambda + beta*yb0 <= yb0
    for k in range(YB.shape[1]):
        row = np.zeros(n + 1); row[:n] = YB[:, k]; row[-1] = yb0[k]
        Aub.append(row); bub.append(yb0[k])
    Aeq = None; beq = None
    if vrs:
        Aeq = np.zeros((1, n + 1)); Aeq[0, :n] = 1.0
        beq = np.array([1.0])
    bounds = [(0, None)] * n + [(None, None)]
    res = linprog(c, A_ub=np.array(Aub), b_ub=np.array(bub), A_eq=Aeq, b_eq=beq,
                  bounds=bounds, method="highs")
    if not res.success:
        raise RuntimeError(res.message)
    return float(res.x[-1])


def main():
    rows = load_data()
    X = np.array([[r["land"]] for r in rows], dtype=float)
    YG = np.array([[r["population"], r["real_gdp"]] for r in rows], dtype=float)
    YB = np.array([[r["co2"]] for r in rows], dtype=float)
    for idx, r in enumerate(rows):
        r["oe"] = sbm_undesirable(X[idx], YG[idx], YB[idx], X, YG, YB, vrs=False)
        r["pte"] = sbm_undesirable(X[idx], YG[idx], YB[idx], X, YG, YB, vrs=True)
        r["se"] = r["oe"] / r["pte"] if r["pte"] else None

    # Global and contemporaneous directional distances for an auditable GML decomposition.
    by_year_idx = {year: [i for i, r in enumerate(rows) if r["year"] == year] for year in YEARS}
    for idx, r in enumerate(rows):
        own = by_year_idx[r["year"]]
        r["dd_global"] = directional_distance(X[idx], YG[idx], YB[idx], X, YG, YB, vrs=False)
        r["dd_contemporary"] = directional_distance(
            X[idx], YG[idx], YB[idx], X[own], YG[own], YB[own], vrs=False
        )
        r["dd_contemporary_vrs"] = directional_distance(
            X[idx], YG[idx], YB[idx], X[own], YG[own], YB[own], vrs=True
        )

    summaries = []
    for year in YEARS:
        rr = [r for r in rows if r["year"] == year]
        oe = np.array([r["oe"] for r in rr])
        pte = np.array([r["pte"] for r in rr])
        se = np.array([r["se"] for r in rr])
        summaries.append({
            "year": year,
            "oe_mean": float(oe.mean()), "oe_median": float(np.median(oe)),
            "oe_sd_pop": float(oe.std(ddof=0)), "oe_sd_sample": float(oe.std(ddof=1)),
            "oe_cv_pop": float(oe.std(ddof=0) / oe.mean()),
            "oe_cv_sample": float(oe.std(ddof=1) / oe.mean()),
            "pte_mean": float(pte.mean()), "se_mean": float(se.mean()),
        })

    convergence = []
    mobility = []
    gml_summary = []
    gml_rows = []
    for y0, y1 in zip(YEARS[:-1], YEARS[1:]):
        a = sorted([r for r in rows if r["year"] == y0], key=lambda z: z["city"])
        b = sorted([r for r in rows if r["year"] == y1], key=lambda z: z["city"])
        assert [r["city"] for r in a] == [r["city"] for r in b]
        e0 = np.array([r["oe"] for r in a])
        e1 = np.array([r["oe"] for r in b])
        annual_growth = np.log(e1 / e0) / (y1 - y0)
        lr = linregress(np.log(e0), annual_growth)
        convergence.append({
            "period": f"{y0}-{y1}", "beta": float(lr.slope), "intercept": float(lr.intercept),
            "se": float(lr.stderr), "t": float(lr.slope / lr.stderr), "p": float(lr.pvalue),
            "r2": float(lr.rvalue ** 2), "n": len(e0),
        })
        rank0 = np.argsort(np.argsort(-e0)) + 1
        rank1 = np.argsort(np.argsort(-e1)) + 1
        rho, p = spearmanr(rank0, rank1)
        mobility.append({
            "period": f"{y0}-{y1}", "spearman_rho": float(rho), "spearman_p": float(p),
            "mean_abs_rank_change": float(np.mean(np.abs(rank1-rank0))),
            "median_abs_rank_change": float(np.median(np.abs(rank1-rank0))),
            "max_abs_rank_change": int(np.max(np.abs(rank1-rank0))),
        })
        for r, rank in zip(a, rank0): r["rank"] = int(rank)
        for r, rank in zip(b, rank1): r["rank"] = int(rank)

        gml_city = []
        for ra, rb in zip(a, b):
            gml = (1.0 + ra["dd_global"]) / (1.0 + rb["dd_global"])
            ec = (1.0 + ra["dd_contemporary"]) / (1.0 + rb["dd_contemporary"])
            pec = (1.0 + ra["dd_contemporary_vrs"]) / (1.0 + rb["dd_contemporary_vrs"])
            sec = ec / pec
            tc = gml / ec
            gml_city.append((gml, tc, ec, pec, sec))
            gml_rows.append({"city": ra["city"], "abbreviation": ra["short"],
                             "period": f"{y0}-{y1}",
                             "gml": gml, "tc": tc, "ec": ec,
                             "pec": pec, "sec": sec})
        gm = lambda vals: float(np.exp(np.mean(np.log(np.array(vals)))))
        gml_summary.append({
            "period": f"{y0}-{y1}",
            "gml_geomean": gm([z[0] for z in gml_city]),
            "tc_geomean": gm([z[1] for z in gml_city]),
            "ec_geomean": gm([z[2] for z in gml_city]),
            "pec_geomean": gm([z[3] for z in gml_city]),
            "sec_geomean": gm([z[4] for z in gml_city]),
            "gml_improve_count": int(sum(z[0] > 1.0 + 1e-9 for z in gml_city)),
            "gml_decline_count": int(sum(z[0] < 1.0 - 1e-9 for z in gml_city)),
            "tc_dominant_count": int(sum(z[1] > z[2] for z in gml_city)),
            "ec_dominant_count": int(sum(z[2] >= z[1] for z in gml_city)),
        })
    # Ensure ranks are assigned for all years independently.
    for year in YEARS:
        rr = [r for r in rows if r["year"] == year]
        vals = np.array([r["oe"] for r in rr])
        ranks = np.argsort(np.argsort(-vals)) + 1
        for r, rank in zip(rr, ranks): r["rank"] = int(rank)

    result = {"rows": rows, "summary": summaries, "convergence": convergence,
              "mobility": mobility, "gml": gml_summary, "gml_rows": gml_rows}
    out = ROOT / "results" / "computed_metrics.json"
    out.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"summary": summaries, "convergence": convergence,
                      "mobility": mobility, "gml": gml_summary}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
